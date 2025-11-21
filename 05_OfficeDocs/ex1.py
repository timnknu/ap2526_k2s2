
import openpyxl

wb = openpyxl.load_workbook('sample_src_table.xlsx')
#print(wb.sheetnames)
wsheet = wb[wb.sheetnames[0]]

for i in range(3, wsheet.max_row+1):
    x = wsheet.cell(row=i, column=1).value
    wsheet.cell(row=i, column=2).value = x**2
    xmin = 5
    xmax = 30
    shade = round( 255 - (x - xmin)/(xmax - xmin) * 255 )
    fill_color = 'FFFF' + f'{shade:02X}'
    print(fill_color)
    wsheet.cell(row=i, column=1).fill = openpyxl.styles.PatternFill(start_color=fill_color,
                end_color=fill_color, fill_type="solid")

wb.save('sample_edited1.xlsx')
